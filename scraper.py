"""
Google Places Business Scraper
==============================
Scrapes businesses in any location using the Google Places API.

Can be used as:
  1. CLI tool:  python scraper.py --location "Gili Air"
  2. Imported:  from scraper import run_scraper, export_to_excel
"""

import argparse
import csv
import json
import math
import os
import re
import sys
import time
from collections import Counter
from dataclasses import dataclass, asdict
from typing import Optional, Callable

import requests

# ─── Configuration Defaults ──────────────────────────────────────────────────

DEFAULT_RADIUS = 3000
REQUEST_DELAY = 0.2

CATEGORY_PRESETS = {
    # ── PRIMARY (always visible) ──────────────────────────────────────────
    "restaurant": [
        {"query": "restaurant", "category": "Restaurant", "type": "restaurant"},
        {"query": "cafe", "category": "Restaurant", "type": "cafe"},
        {"query": "bar", "category": "Restaurant", "type": "bar"},
    ],
    "coffee-shops": [
        {"query": "coffee shop", "category": "Coffee Shops", "type": "cafe"},
        {"query": "cafe", "category": "Coffee Shops", "type": "cafe"},
    ],
    "hotel": [
        {"query": "hotel", "category": "Hotel", "type": "lodging"},
        {"query": "resort", "category": "Hotel", "type": None},
    ],
    "guest-house": [
        {"query": "guest house", "category": "Guest House", "type": None},
        {"query": "homestay", "category": "Guest House", "type": None},
        {"query": "hostel", "category": "Guest House", "type": None},
        {"query": "villa", "category": "Guest House", "type": None},
        {"query": "bungalow", "category": "Guest House", "type": None},
    ],
    "things-to-do": [
        {"query": "things to do", "category": "Things To Do", "type": None},
        {"query": "tourist attraction", "category": "Things To Do", "type": "tourist_attraction"},
    ],
    "nightlife": [
        {"query": "nightclub", "category": "Nightlife", "type": "night_club"},
        {"query": "bar", "category": "Nightlife", "type": "bar"},
        {"query": "lounge", "category": "Nightlife", "type": None},
    ],
    "spa-beauty": [
        {"query": "spa", "category": "Spa & Beauty", "type": "spa"},
        {"query": "beauty salon", "category": "Spa & Beauty", "type": "beauty_salon"},
        {"query": "hair salon", "category": "Spa & Beauty", "type": "hair_care"},
    ],
    "gym-fitness": [
        {"query": "gym", "category": "Gym & Fitness", "type": "gym"},
        {"query": "fitness center", "category": "Gym & Fitness", "type": None},
    ],
    "shopping": [
        {"query": "shopping", "category": "Shopping", "type": "shopping_mall"},
        {"query": "market", "category": "Shopping", "type": None},
        {"query": "store", "category": "Shopping", "type": "store"},
    ],
    "health": [
        {"query": "hospital", "category": "Health & Medical", "type": "hospital"},
        {"query": "clinic", "category": "Health & Medical", "type": None},
        {"query": "pharmacy", "category": "Health & Medical", "type": "pharmacy"},
    ],
    "real-estate": [
        {"query": "real estate agency", "category": "Real Estate", "type": "real_estate_agency"},
        {"query": "real estate agent", "category": "Real Estate", "type": "real_estate_agency"},
        {"query": "property management", "category": "Real Estate", "type": None},
    ],
    "coworking": [
        {"query": "coworking space", "category": "Coworking", "type": None},
        {"query": "coworking", "category": "Coworking", "type": None},
    ],
    # ── SECONDARY (show more) ─────────────────────────────────────────────
    "dentist": [
        {"query": "dentist", "category": "Dentist", "type": "dentist"},
        {"query": "dental clinic", "category": "Dentist", "type": "dentist"},
    ],
    "veterinary": [
        {"query": "veterinarian", "category": "Veterinary", "type": "veterinary_care"},
        {"query": "animal hospital", "category": "Veterinary", "type": "veterinary_care"},
    ],
    "car-services": [
        {"query": "car repair", "category": "Car Services", "type": "car_repair"},
        {"query": "car wash", "category": "Car Services", "type": "car_wash"},
        {"query": "car rental", "category": "Car Services", "type": "car_rental"},
    ],
    "supermarket": [
        {"query": "supermarket", "category": "Supermarket", "type": "supermarket"},
        {"query": "grocery store", "category": "Supermarket", "type": "convenience_store"},
    ],
    "bank": [
        {"query": "bank", "category": "Bank & ATM", "type": "bank"},
        {"query": "atm", "category": "Bank & ATM", "type": "atm"},
    ],
    "pharmacy": [
        {"query": "pharmacy", "category": "Pharmacy", "type": "pharmacy"},
        {"query": "drugstore", "category": "Pharmacy", "type": None},
    ],
    "education": [
        {"query": "school", "category": "School & Education", "type": "school"},
        {"query": "university", "category": "School & Education", "type": "university"},
        {"query": "language school", "category": "School & Education", "type": None},
    ],
    "yoga-pilates": [
        {"query": "yoga studio", "category": "Yoga & Pilates", "type": None},
        {"query": "pilates studio", "category": "Yoga & Pilates", "type": None},
    ],
    "laundry": [
        {"query": "laundry", "category": "Laundry", "type": "laundry"},
        {"query": "dry cleaning", "category": "Laundry", "type": None},
    ],
    "pet-store": [
        {"query": "pet store", "category": "Pet Store", "type": "pet_store"},
        {"query": "pet shop", "category": "Pet Store", "type": "pet_store"},
    ],
    "electronics": [
        {"query": "electronics store", "category": "Electronics", "type": "electronics_store"},
        {"query": "phone repair", "category": "Electronics", "type": None},
    ],
    "furniture-home": [
        {"query": "furniture store", "category": "Furniture & Home", "type": "furniture_store"},
        {"query": "home goods store", "category": "Furniture & Home", "type": "home_goods_store"},
    ],
    "travel-agency": [
        {"query": "travel agency", "category": "Travel Agency", "type": "travel_agency"},
        {"query": "tour operator", "category": "Travel Agency", "type": None},
    ],
    "barbershop": [
        {"query": "barbershop", "category": "Barbershop", "type": "hair_care"},
        {"query": "barber", "category": "Barbershop", "type": "hair_care"},
    ],
    "physiotherapy": [
        {"query": "physiotherapist", "category": "Physiotherapy", "type": "physiotherapist"},
        {"query": "physical therapy", "category": "Physiotherapy", "type": "physiotherapist"},
    ],
}

PRIMARY_CATEGORIES = [
    "restaurant", "coffee-shops", "hotel", "guest-house", "things-to-do",
    "nightlife", "spa-beauty", "gym-fitness", "shopping", "health",
    "real-estate", "coworking",
]

SECONDARY_CATEGORIES = [
    "dentist", "veterinary", "car-services", "supermarket", "bank",
    "pharmacy", "education", "yoga-pilates", "laundry", "pet-store",
    "electronics", "furniture-home", "travel-agency", "barbershop",
    "physiotherapy",
]

DEFAULT_CATEGORIES = ["restaurant", "coffee-shops", "hotel", "things-to-do", "spa-beauty"]

# Maps Google Place types → our category labels.
# Used to re-categorize businesses based on what Google actually classifies them as.
GOOGLE_TYPE_TO_CATEGORY = {
    "restaurant": "Restaurant",
    "cafe": "Restaurant",
    "bar": "Restaurant",
    "food": "Restaurant",
    "bakery": "Restaurant",
    "meal_delivery": "Restaurant",
    "meal_takeaway": "Restaurant",
    "tourist_attraction": "Things To Do",
    "amusement_park": "Things To Do",
    "aquarium": "Things To Do",
    "art_gallery": "Things To Do",
    "museum": "Things To Do",
    "park": "Things To Do",
    "zoo": "Things To Do",
    "spa": "Spa & Beauty",
    "beauty_salon": "Spa & Beauty",
    "hair_care": "Spa & Beauty",
    "lodging": "Hotel",
    "night_club": "Nightlife",
    "gym": "Gym & Fitness",
    "shopping_mall": "Shopping",
    "store": "Shopping",
    "clothing_store": "Shopping",
    "supermarket": "Supermarket",
    "convenience_store": "Supermarket",
    "hospital": "Health & Medical",
    "pharmacy": "Pharmacy",
    "doctor": "Health & Medical",
    "dentist": "Dentist",
    "health": "Health & Medical",
    "drugstore": "Pharmacy",
    "real_estate_agency": "Real Estate",
    "car_repair": "Car Services",
    "car_wash": "Car Services",
    "car_rental": "Car Services",
    "car_dealer": "Car Services",
    "bank": "Bank & ATM",
    "atm": "Bank & ATM",
    "school": "School & Education",
    "university": "School & Education",
    "primary_school": "School & Education",
    "secondary_school": "School & Education",
    "veterinary_care": "Veterinary",
    "laundry": "Laundry",
    "pet_store": "Pet Store",
    "electronics_store": "Electronics",
    "furniture_store": "Furniture & Home",
    "home_goods_store": "Furniture & Home",
    "travel_agency": "Travel Agency",
    "physiotherapist": "Physiotherapy",
}

# Google types that should EXCLUDE a business from a given category.
# e.g. if searching for "Gym & Fitness", exclude places Google classifies as yoga/spa.
CATEGORY_EXCLUDE_TYPES = {
    "Gym & Fitness": {"spa", "beauty_salon", "hair_care", "yoga_studio", "travel_agency",
                      "lodging", "restaurant", "cafe", "bar", "food"},
    "Spa & Beauty": {"gym", "restaurant", "cafe", "bar", "food", "lodging"},
    "Restaurant": {"lodging", "gym", "spa"},
    "Hotel": {"restaurant", "cafe", "bar", "food", "gym", "spa"},
    "Guest House": {"restaurant", "cafe", "bar", "food", "gym", "spa"},
    "Yoga & Pilates": {"gym", "restaurant", "cafe", "bar", "lodging"},
    "Barbershop": {"spa", "beauty_salon", "restaurant", "cafe"},
}


# ─── Data Model ──────────────────────────────────────────────────────────────

@dataclass
class Business:
    name: str = ""
    category: str = ""
    google_types: str = ""
    address: str = ""
    phone: str = ""
    website: str = ""
    rating: Optional[float] = None
    reviews_count: Optional[int] = None
    price_level: str = ""
    opening_hours: str = ""
    description: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    google_maps_url: str = ""
    photo_url: str = ""
    place_id: str = ""


# ─── API Functions ───────────────────────────────────────────────────────────

def geocode_location(location_name: str, api_key: str) -> tuple:
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": location_name, "key": api_key}
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()
    data = response.json()

    if data.get("status") != "OK" or not data.get("results"):
        status = data.get("status", "UNKNOWN")
        error_msg = data.get("error_message", "")
        if status == "REQUEST_DENIED":
            raise ValueError(
                f"Geocoding API request denied. Please enable the 'Geocoding API' "
                f"in your Google Cloud Console (console.cloud.google.com/apis/library). "
                f"You need both 'Places API' and 'Geocoding API' enabled. {error_msg}"
            )
        raise ValueError(f"Could not geocode '{location_name}'. Status: {status}. {error_msg}")

    result = data["results"][0]
    loc = result["geometry"]["location"]
    formatted = result.get("formatted_address", location_name)
    return loc["lat"], loc["lng"], formatted


def _text_search(query: str, location_name: str, lat: float, lng: float,
                 radius: int, api_key: str, page_token: str = None) -> dict:
    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {
        "query": f"{query} in {location_name}",
        "location": f"{lat},{lng}",
        "radius": radius,
        "key": api_key,
    }
    if page_token:
        params["pagetoken"] = page_token
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()
    return response.json()


def _get_place_details(place_id: str, api_key: str) -> dict:
    url = "https://maps.googleapis.com/maps/api/place/details/json"
    params = {
        "place_id": place_id,
        "fields": (
            "name,formatted_address,formatted_phone_number,international_phone_number,"
            "website,rating,user_ratings_total,price_level,opening_hours,"
            "editorial_summary,geometry,url,types,business_status,photos"
        ),
        "key": api_key,
    }
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()
    return response.json()


def _nearby_search(lat: float, lng: float, radius: int,
                   place_type: str, api_key: str, page_token: str = None) -> dict:
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": radius,
        "type": place_type,
        "key": api_key,
    }
    if page_token:
        params["pagetoken"] = page_token
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()
    return response.json()


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _is_within_radius(lat, lng, center_lat, center_lng, radius_m):
    R = 6371000
    phi1 = math.radians(center_lat)
    phi2 = math.radians(lat)
    dphi = math.radians(lat - center_lat)
    dlam = math.radians(lng - center_lng)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c <= radius_m * 1.2


def _format_opening_hours(hours_data):
    if not hours_data:
        return ""
    weekday_text = hours_data.get("weekday_text", [])
    return " | ".join(weekday_text) if weekday_text else ""


def _price_level_to_string(level):
    if level is None:
        return ""
    return "$" * level if level > 0 else "Free"


def _resolve_category(google_types: list, fallback_category: str) -> str:
    """Determine the best category label from Google's actual place types.

    Prefers a matching GOOGLE_TYPE_TO_CATEGORY entry over the search-preset label.
    Falls back to the preset label if no specific match is found.
    """
    for gtype in google_types:
        mapped = GOOGLE_TYPE_TO_CATEGORY.get(gtype)
        if mapped:
            return mapped
    return fallback_category


def _get_expected_types(category: str) -> set:
    """Return the set of Google types that legitimately belong to a category."""
    return {gtype for gtype, cat in GOOGLE_TYPE_TO_CATEGORY.items() if cat == category}


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '_', text)
    return text


def resolve_search_queries(category_input) -> list:
    """Convert category names (list or comma-string) into search query dicts."""
    if isinstance(category_input, str):
        categories = [c.strip() for c in category_input.split(",") if c.strip()]
    else:
        categories = list(category_input)

    queries = []
    for cat in categories:
        cat_key = cat.lower().replace(" ", "-")
        if cat_key in CATEGORY_PRESETS:
            queries.extend(CATEGORY_PRESETS[cat_key])
        else:
            nice_name = cat.strip().title()
            queries.append({"query": cat.strip(), "category": nice_name, "type": None})
    return queries


# ─── Core Scraper (importable) ───────────────────────────────────────────────

def run_scraper(
    api_key: str,
    location: str,
    categories: list,
    radius: int = DEFAULT_RADIUS,
    lat: float = None,
    lng: float = None,
    progress_callback: Callable = None,
) -> list:
    """
    Main scraper function. Can be called from web app or CLI.

    Args:
        api_key: Google Places API key
        location: Location name (e.g. "Gili Air")
        categories: List of category preset names or free-text queries
        radius: Search radius in meters
        lat/lng: Optional coordinates (auto-geocoded if not provided)
        progress_callback: Optional fn(message: str, percent: int) for live updates

    Returns:
        List of Business dataclass objects
    """

    def emit(msg, pct=None):
        if progress_callback:
            progress_callback(msg, pct)

    # Resolve categories to search queries
    search_queries = resolve_search_queries(categories)
    if not search_queries:
        raise ValueError("No categories selected")

    # Geocode if needed
    if lat is None or lng is None:
        emit(f"Geocoding '{location}'...", 2)
        lat, lng, formatted = geocode_location(location, api_key)
        emit(f"Found: {formatted} ({lat:.4f}, {lng:.4f})", 5)
    else:
        formatted = location

    emit(f"Searching {len(search_queries)} queries in {formatted} (radius: {radius}m)", 5)

    # ── Step 1: Collect place IDs ──
    place_ids = {}
    total_queries = len(search_queries)

    for qi, sq in enumerate(search_queries):
        query = sq["query"]
        category = sq["category"]
        place_type = sq.get("type")

        base_pct = 5 + int((qi / total_queries) * 40)  # 5% → 45%
        emit(f"Searching: '{query}' ({category})", base_pct)

        # Text search with pagination
        try:
            result = _text_search(query, location, lat, lng, radius, api_key)
        except requests.RequestException as e:
            emit(f"Text search failed for '{query}': {e}")
            continue

        status = result.get("status", "UNKNOWN")
        if status == "REQUEST_DENIED":
            raise ValueError(f"API key denied. Check that your key is valid and Places API is enabled. {result.get('error_message', '')}")
        if status not in ("OK", "ZERO_RESULTS"):
            emit(f"API status for '{query}': {status}")
            continue

        # Process pages
        page_count = 0
        while True:
            page_count += 1
            results = result.get("results", [])

            for place in results:
                pid = place.get("place_id")
                if pid and pid not in place_ids:
                    loc = place.get("geometry", {}).get("location", {})
                    p_lat, p_lng = loc.get("lat", 0), loc.get("lng", 0)
                    if _is_within_radius(p_lat, p_lng, lat, lng, radius):
                        place_ids[pid] = category

            next_token = result.get("next_page_token")
            if not next_token or page_count >= 3:
                break
            time.sleep(2)
            try:
                result = _text_search(query, location, lat, lng, radius, api_key, page_token=next_token)
            except requests.RequestException:
                break

        # Nearby search if type available
        if place_type:
            time.sleep(REQUEST_DELAY)
            try:
                result = _nearby_search(lat, lng, radius, place_type, api_key)
                nb_page = 0
                while True:
                    nb_page += 1
                    for place in result.get("results", []):
                        pid = place.get("place_id")
                        if pid and pid not in place_ids:
                            loc = place.get("geometry", {}).get("location", {})
                            p_lat, p_lng = loc.get("lat", 0), loc.get("lng", 0)
                            if _is_within_radius(p_lat, p_lng, lat, lng, radius):
                                place_ids[pid] = category

                    next_token = result.get("next_page_token")
                    if not next_token or nb_page >= 3:
                        break
                    time.sleep(2)
                    try:
                        result = _nearby_search(lat, lng, radius, place_type, api_key, page_token=next_token)
                    except requests.RequestException:
                        break
            except requests.RequestException:
                pass

        time.sleep(REQUEST_DELAY)

    emit(f"Found {len(place_ids)} unique places. Fetching details...", 45)

    if not place_ids:
        emit("No places found. Try a larger radius or different categories.", 100)
        return []

    # ── Step 2: Fetch place details ──
    businesses = []
    total_places = len(place_ids)

    for i, (place_id, category) in enumerate(place_ids.items(), 1):
        pct = 45 + int((i / total_places) * 50)  # 45% → 95%
        emit(f"[{i}/{total_places}] Fetching details...", pct)

        try:
            result = _get_place_details(place_id, api_key)
        except requests.RequestException as e:
            emit(f"Failed to fetch details: {e}")
            continue

        if result.get("status") != "OK":
            continue

        place = result.get("result", {})

        if place.get("business_status") == "CLOSED_PERMANENTLY":
            emit(f"Skipped (closed): {place.get('name', '?')}")
            continue

        google_types = place.get("types", [])

        # Determine the real category from Google's types
        real_category = _resolve_category(google_types, category)

        # Filter out mismatches: if Google's types indicate this place
        # doesn't belong in the requested category, skip it
        exclude_types = CATEGORY_EXCLUDE_TYPES.get(category, set())
        if exclude_types and any(t in exclude_types for t in google_types):
            # Only skip if the place has NO types matching the requested category
            requested_types = _get_expected_types(category)
            if not any(t in requested_types for t in google_types):
                emit(f"Skipped (mismatch): {place.get('name', '?')} — Google types: {', '.join(google_types)}")
                continue

        loc = place.get("geometry", {}).get("location", {})

        # Build photo URL from first photo reference
        photo_url = ""
        photos = place.get("photos", [])
        if photos:
            ref = photos[0].get("photo_reference", "")
            if ref:
                photo_url = (
                    f"https://maps.googleapis.com/maps/api/place/photo"
                    f"?maxwidth=400&photo_reference={ref}&key={api_key}"
                )

        biz = Business(
            name=place.get("name", ""),
            category=real_category,
            google_types=", ".join(google_types),
            address=place.get("formatted_address", ""),
            phone=place.get("international_phone_number", "") or place.get("formatted_phone_number", ""),
            website=place.get("website", ""),
            rating=place.get("rating"),
            reviews_count=place.get("user_ratings_total"),
            price_level=_price_level_to_string(place.get("price_level")),
            opening_hours=_format_opening_hours(place.get("opening_hours")),
            description=(place.get("editorial_summary") or {}).get("overview", ""),
            latitude=loc.get("lat"),
            longitude=loc.get("lng"),
            google_maps_url=place.get("url", ""),
            photo_url=photo_url,
            place_id=place_id,
        )
        businesses.append(biz)
        emit(f"[{i}/{total_places}] {biz.name}", pct)

        time.sleep(REQUEST_DELAY)

    businesses.sort(key=lambda b: (b.category, -(b.rating or 0)))
    emit(f"Done! Found {len(businesses)} businesses.", 100)
    return businesses


# ─── Export Functions ────────────────────────────────────────────────────────

def export_to_excel(businesses: list, filepath: str):
    """Export businesses to a formatted .xlsx Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Businesses"

    # Column config
    columns = [
        ("Name", 35),
        ("Category", 16),
        ("Rating", 10),
        ("Reviews", 10),
        ("Price", 8),
        ("Phone", 20),
        ("Website", 35),
        ("Address", 45),
        ("Opening Hours", 60),
        ("Description", 50),
        ("Google Types", 30),
        ("Latitude", 12),
        ("Longitude", 12),
        ("Google Maps URL", 40),
        ("Photo URL", 45),
        ("Place ID", 30),
    ]

    # Header styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = col_width

    # Data styling
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Write data
    field_order = [
        "name", "category", "rating", "reviews_count", "price_level",
        "phone", "website", "address", "opening_hours", "description",
        "google_types", "latitude", "longitude", "google_maps_url", "photo_url", "place_id",
    ]

    for row_idx, biz in enumerate(businesses, 2):
        biz_dict = asdict(biz)
        for col_idx, field in enumerate(field_order, 1):
            value = biz_dict.get(field, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    counts = Counter(b.category for b in businesses)
    ws2.cell(row=1, column=1, value="Category").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Count").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill

    for i, (cat, count) in enumerate(sorted(counts.items()), 2):
        ws2.cell(row=i, column=1, value=cat).font = data_font
        ws2.cell(row=i, column=2, value=count).font = data_font

    total_row = len(counts) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=total_row, column=2, value=len(businesses)).font = Font(name="Arial", bold=True, size=10)

    rated = [b for b in businesses if b.rating is not None]
    if rated:
        avg_row = total_row + 2
        avg_rating = sum(b.rating for b in rated) / len(rated)
        ws2.cell(row=avg_row, column=1, value="Average Rating").font = data_font
        ws2.cell(row=avg_row, column=2, value=round(avg_rating, 1)).font = data_font

    wb.save(filepath)


def export_to_csv(businesses: list, filepath: str):
    """Export businesses to CSV."""
    if not businesses:
        return

    fieldnames = [
        "name", "category", "rating", "reviews_count", "price_level",
        "phone", "website", "address", "opening_hours", "description",
        "google_types", "latitude", "longitude", "google_maps_url", "photo_url", "place_id",
    ]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for biz in businesses:
            writer.writerow(asdict(biz))


def get_summary(businesses: list) -> dict:
    """Return a summary dict for the web UI."""
    counts = dict(Counter(b.category for b in businesses))
    rated = [b for b in businesses if b.rating is not None]
    avg_rating = round(sum(b.rating for b in rated) / len(rated), 1) if rated else None

    top5 = sorted(rated, key=lambda b: (-b.rating, -(b.reviews_count or 0)))[:5]
    top_list = [
        {"name": b.name, "category": b.category, "rating": b.rating, "reviews": b.reviews_count or 0}
        for b in top5
    ]

    # Full business data for preview table and map
    biz_data = [
        {
            "name": b.name, "category": b.category, "rating": b.rating,
            "reviews_count": b.reviews_count or 0, "phone": b.phone,
            "website": b.website, "address": b.address,
            "latitude": b.latitude, "longitude": b.longitude,
            "photo_url": b.photo_url, "google_maps_url": b.google_maps_url,
            "price_level": b.price_level,
        }
        for b in businesses
    ]

    return {
        "total": len(businesses),
        "by_category": counts,
        "avg_rating": avg_rating,
        "rated_count": len(rated),
        "top5": top_list,
        "businesses": biz_data,
    }


# ─── CLI (kept for standalone usage) ────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Scrape businesses from Google Maps.")
    parser.add_argument("--location", "-l", required=True)
    parser.add_argument("--lat", type=float)
    parser.add_argument("--lng", type=float)
    parser.add_argument("--radius", "-r", type=int, default=DEFAULT_RADIUS)
    parser.add_argument("--categories", "-c", default=",".join(DEFAULT_CATEGORIES))
    parser.add_argument("--output", "-o")
    parser.add_argument("--list-categories", action="store_true")

    args = parser.parse_args()

    if args.list_categories:
        for name, queries in sorted(CATEGORY_PRESETS.items()):
            terms = ", ".join(q["query"] for q in queries)
            print(f"  {name:<16} -> {terms}")
        return

    api_key = os.environ.get("GOOGLE_PLACES_API_KEY", "")
    if not api_key:
        print("Set GOOGLE_PLACES_API_KEY environment variable.")
        sys.exit(1)

    categories = [c.strip() for c in args.categories.split(",")]

    def cli_progress(msg, pct=None):
        prefix = f"[{pct}%] " if pct is not None else ""
        print(f"{prefix}{msg}")

    businesses = run_scraper(
        api_key=api_key,
        location=args.location,
        categories=categories,
        radius=args.radius,
        lat=args.lat,
        lng=args.lng,
        progress_callback=cli_progress,
    )

    if businesses:
        output = args.output or f"{slugify(args.location)}_businesses.csv"
        export_to_csv(businesses, output)
        print(f"\nExported {len(businesses)} businesses to {output}")


if __name__ == "__main__":
    main()
